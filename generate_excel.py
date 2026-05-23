from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import date

BLUE   = "2D62C8"
WHITE  = "FFFFFF"
LIGHT  = "EEF2FB"
YELLOW = "FFF2CC"
GREEN  = "D9EAD3"
ORANGE = "FCE5CD"
GRAY   = "F3F3F3"

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def bold_font(size=11, color="000000", italic=False):
    return Font(bold=True, size=size, color=color, italic=italic)

def normal_font(size=10):
    return Font(size=size)

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def apply_table_style(ws, data_rows, start_row, cols):
    for row_idx, row_data in enumerate(data_rows):
        r = start_row + row_idx
        fill = PatternFill("solid", fgColor=LIGHT if row_idx % 2 == 0 else WHITE)
        for c, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.fill = fill
            cell.font = normal_font()
            cell.border = thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

wb = Workbook()

# ── Sheet 1: Overview ────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Overview"

ws1.merge_cells("A1:E1")
title = ws1["A1"]
title.value = f"Impedyme Competitor Analysis — {date.today().strftime('%Y-%m-%d')}"
title.font = bold_font(16, WHITE)
title.fill = header_fill(BLUE)
title.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:E2")
sub = ws1["A2"]
sub.value = "HIL & Real-Time Simulation Market — Competitive Intelligence Report"
sub.font = Font(size=11, italic=True, color="555555")
sub.alignment = Alignment(horizontal="center")
ws1.row_dimensions[2].height = 22

headers = ["Competitor", "Headquarters", "Core Focus", "Key Differentiator", "Threat Level"]
for c, h in enumerate(headers, start=1):
    cell = ws1.cell(row=4, column=c, value=h)
    cell.font = bold_font(11, WHITE)
    cell.fill = header_fill(BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
ws1.row_dimensions[4].height = 24

overview_data = [
    ["Speedgoat", "Natick, MA / Liebefeld, CH", "Simulink real-time target machines", "Deep MathWorks integration; official Simulink partner", "High"],
    ["Typhoon HIL", "Novi, MI / Kragujevac, RS", "Compact HIL simulators & grid emulators", "Fastest time-step on the market; strong power-electronics focus", "High"],
    ["OPAL-RT", "Montreal, QC, CA", "Real-time digital simulators for power systems", "Decades of power-grid heritage; large installed base in utilities", "Medium"],
]
apply_table_style(ws1, overview_data, 5, 5)

# Colour-code threat level column
threat_colors = {"High": "F4CCCC", "Medium": YELLOW, "Low": GREEN}
for row_idx, row in enumerate(overview_data):
    cell = ws1.cell(row=5 + row_idx, column=5)
    cell.fill = PatternFill("solid", fgColor=threat_colors.get(row[4], WHITE))
    cell.font = bold_font(10)
    cell.alignment = Alignment(horizontal="center", vertical="top")

set_col_widths(ws1, {"A": 18, "B": 26, "C": 32, "D": 44, "E": 16})

# ── Sheet 2: Feature Comparison ──────────────────────────────────────────────
ws2 = wb.create_sheet("Feature Comparison")

ws2.merge_cells("A1:D1")
h = ws2["A1"]
h.value = "Feature Comparison Matrix"
h.font = bold_font(14, WHITE)
h.fill = header_fill(BLUE)
h.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

col_headers = ["Feature / Capability", "Speedgoat", "Typhoon HIL", "OPAL-RT"]
for c, txt in enumerate(col_headers, start=1):
    cell = ws2.cell(row=3, column=c, value=txt)
    cell.font = bold_font(11, WHITE)
    cell.fill = header_fill(BLUE)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border()

feature_data = [
    ["Simulink / MATLAB Integration", "Native (official partner)", "Good (third-party)", "Good (third-party)"],
    ["Minimum Time Step", "~1 µs", "< 1 µs (FPGA)", "~1–2 µs"],
    ["FPGA-Based Simulation", "Yes (Speedgoat FPGA I/O)", "Yes (core strength)", "Yes (OP5707 series)"],
    ["Power Electronics / Grid Focus", "Moderate", "Very Strong", "Very Strong"],
    ["Automotive / ADAS Support", "Strong", "Moderate", "Moderate"],
    ["Cloud / Remote Access", "Limited", "Typhoon Cloud", "RT-LAB Cloud"],
    ["Pricing Model", "Hardware + licence", "Hardware + licence", "Hardware + licence"],
    ["Open API / SDK", "Simulink-centric", "Python API", "eHS SDK, RT-LAB API"],
    ["Training & Certification", "MathWorks ecosystem", "HIL Academy", "OPAL-RT Academy"],
    ["Support Model", "Regional partners", "Direct + partners", "Direct + partners"],
]
apply_table_style(ws2, feature_data, 4, 4)
set_col_widths(ws2, {"A": 36, "B": 30, "C": 30, "D": 30})

# ── Sheet 3: Recent Updates ───────────────────────────────────────────────────
ws3 = wb.create_sheet("Recent Updates")

ws3.merge_cells("A1:E1")
h = ws3["A1"]
h.value = "Recent Competitor Updates & News"
h.font = bold_font(14, WHITE)
h.fill = header_fill(BLUE)
h.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

col_headers = ["Competitor", "Category", "Title", "Summary", "Significance"]
for c, txt in enumerate(col_headers, start=1):
    cell = ws3.cell(row=3, column=c, value=txt)
    cell.font = bold_font(11, WHITE)
    cell.fill = header_fill(BLUE)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border()

updates_data = [
    ["Speedgoat", "Product Update", "New FPGA I/O Module Gen3", "Next-gen FPGA board with 2× channel density and lower latency for Simulink real-time testing.", "High"],
    ["Speedgoat", "Blog Post", "Model-Based Design for EV Battery BMS", "Case study showing HIL validation of a battery management system using Speedgoat hardware.", "Medium"],
    ["Typhoon HIL", "Product Update", "Typhoon HIL 604+", "Updated compact HIL unit with enhanced grid-forming inverter emulation and IEC 61850 support.", "High"],
    ["Typhoon HIL", "Blog Post", "HIL Testing for Grid-Scale BESS", "Technical article on validating large-scale battery energy storage controllers with HIL.", "Medium"],
    ["Typhoon HIL", "Product Update", "Python API v3.0 Release", "Major API overhaul enabling headless test automation and CI/CD pipeline integration.", "High"],
    ["OPAL-RT", "Product Update", "OP5707XG Real-Time Simulator", "New high-channel-count simulator targeting large power grid and microgrid simulations.", "Medium"],
    ["OPAL-RT", "Blog Post", "Real-Time Simulation for Digital Twins", "White paper on using RT-LAB for digital twin deployment in smart-grid projects.", "Medium"],
    ["OPAL-RT", "Partnership", "Integration with Siemens PSS®E", "Announced interoperability between RT-LAB and Siemens power-flow solver for utility-scale testing.", "High"],
]
apply_table_style(ws3, updates_data, 4, 5)

sig_colors = {"High": "F4CCCC", "Medium": YELLOW, "Low": GREEN}
for row_idx, row in enumerate(updates_data):
    cell = ws3.cell(row=4 + row_idx, column=5)
    cell.fill = PatternFill("solid", fgColor=sig_colors.get(row[4], WHITE))
    cell.font = bold_font(10)
    cell.alignment = Alignment(horizontal="center", vertical="top")

set_col_widths(ws3, {"A": 18, "B": 22, "C": 36, "D": 56, "E": 16})

# ── Sheet 4: Strategic Insights ───────────────────────────────────────────────
ws4 = wb.create_sheet("Strategic Insights")

ws4.merge_cells("A1:B1")
h = ws4["A1"]
h.value = "Strategic Insights & Recommendations for Impedyme"
h.font = bold_font(14, WHITE)
h.fill = header_fill(BLUE)
h.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

insight_headers = ["Topic", "Detail"]
for c, txt in enumerate(insight_headers, start=1):
    cell = ws4.cell(row=3, column=c, value=txt)
    cell.font = bold_font(11, WHITE)
    cell.fill = header_fill(BLUE)
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border()

insights = [
    ["FPGA Speed is the New Baseline",
     "Both Typhoon HIL and OPAL-RT now offer sub-microsecond FPGA time-steps. Impedyme should highlight any latency or resolution advantages clearly in positioning."],
    ["Python API as a Buying Signal",
     "Typhoon HIL's Python API v3 targets DevOps-minded test engineers. Impedyme should prioritise automation-friendly interfaces to compete for this segment."],
    ["Power Electronics is a Battleground",
     "All three competitors are intensifying their grid and power-electronics messaging. If Impedyme's roadmap touches EV or renewable energy testing, this is the right moment to invest."],
    ["Cloud/Remote Access is Emerging",
     "Typhoon Cloud and RT-LAB Cloud suggest a market shift. Impedyme can differentiate by offering a more seamless or cost-effective remote simulation workflow."],
    ["Partnerships Amplify Reach",
     "OPAL-RT's Siemens tie-up and Speedgoat's MathWorks partnership show that ecosystem deals are a key growth lever. Impedyme should identify 1–2 strategic integration partners."],
    ["Pricing Transparency as Differentiator",
     "All three competitors use opaque hardware + licence bundles. A more transparent or subscription-based model could lower the buyer's barrier to evaluation."],
]

for row_idx, row in enumerate(insights):
    r = 4 + row_idx
    fill = PatternFill("solid", fgColor=LIGHT if row_idx % 2 == 0 else WHITE)
    for c, val in enumerate(row, start=1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.fill = fill
        cell.font = bold_font(10) if c == 1 else normal_font(10)
        cell.border = thin_border()
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[r].height = 52

set_col_widths(ws4, {"A": 34, "B": 90})

out = "/home/user/impedyme-competitor-analysis/competitor_analysis_report.xlsx"
wb.save(out)
print(f"Saved: {out}")
